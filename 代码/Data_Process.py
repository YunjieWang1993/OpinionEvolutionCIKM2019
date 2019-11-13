import os
from openpyxl import load_workbook
import jieba
import re

def Split_Data(dir):
    workbook = load_workbook(dir)
    sheet = workbook['Sheet1']
    starttime = sheet.cell(row=2, column=4).value
    starttime = int(starttime[0:4]+starttime[5:7]+starttime[8:10])
    endtime = sheet.cell(row=sheet.max_row, column=4).value
    endtime = int(endtime[0:4] + endtime[5:7] + endtime[8:10])
    gap = 2
    timelist = []
    linesbyday = {}
    for num in range(2,sheet.max_row):
        time = sheet.cell(row=num, column=4).value
        time = int(time[0:4]+time[5:7]+time[8:10])
        if time not in timelist:
            timelist.append(time)
    for t in timelist:
        linesbyday[t] = []
        for num in range(2, sheet.max_row):
            time = sheet.cell(row=num, column=4).value
            time = int(time[0:4] + time[5:7] + time[8:10])
            if time == t:
                linesbyday[t].append(sheet.cell(row=num, column=5).value)
    for t in timelist:
        f = open('E:\SIGIR\事件数据含重复微博版本\山东/'+str(t)+'.txt','w',encoding='utf-8')
        for line in linesbyday[t]:
            line = re.sub('[\s+\.\!\/_$%^*(+\"\')a-zA-Z0-9]+|[+——()?【】“”！。？、~@#￥%……&*（）]+', "", line)
            f.write(line+'\n')
        f.close()


if __name__ == "__main__":
    dir = 'E:\SIGIR\事件数据含重复微博版本\山东\weiboData_sdrma.xlsx'
    Split_Data(dir)