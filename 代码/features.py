import sentiment
import jieba
from openpyxl import load_workbook
import jieba
import math

def sentimentavg(line,sentiment_words,sentiment_dict):
    #print(line)
    stop_words = [w.strip() for w in open('D:/success/ciku/chinese_stopword.txt').readlines()]
    word_list = jieba.cut(line)
    word_list = [w for w in word_list if w not in stop_words and len(w) >= 2]
    word_list_2 = []
    for word in word_list:
        if word in sentiment_words and word not in word_list_2:
            word_list_2.append(word)
    sum = 0
    for word in word_list_2:
        #print(word+' '+str(sentiment_dict[word]))
        sum = sum + sentiment_dict[word]
    if len(word_list_2)==0:
        return 0
    else:
        return sum

def findmaxsentiment(event):
    sentiment_dict = {}
    sentiment_words = []
    workbook = load_workbook('E:\SIGIR\SVM/大连理工大学情感词汇本体库.xlsx')
    sheet = workbook['Sheet1']
    for num in range(2, sheet.max_row):
        sentiment_dict[sheet.cell(row=num, column=1).value] = abs(sheet.cell(row=num, column=6).value)
        sentiment_words.append(sheet.cell(row=num, column=1).value)
    lines = open('E:\SIGIR\事件新闻/' + event + '.txt', 'r').readlines()
    sentimentscore = {}
    for line in lines:
        line = line.replace('\n','')
        sentimentscore[line] = sentimentavg(line, sentiment_words, sentiment_dict)
    return max(sentimentscore,key=sentimentscore.get)

def judgeoneline(event,word,triadsets):
    lines = open('E:\SIGIR\SVM\特征\反转/' + event + '/' + event + '.txt', 'r').readlines()
    triadlist = []
    for item in triadsets:
        if item[2] == word:
            triadlist.append(item)
    judgelist = []
    for item in triadlist:
        judgelist.append(item[0])
        judgelist.append(item[1])
        judgelist.append(item[2])
    resultlist = []
    for line in lines:
        isresult = True
        for i in judgelist:
            if i not in line:
                isresult = False
        if isresult == True:
            resultlist.append(line)
    #print(resultlist)
    if len(resultlist) == 1:
        return False
    else:
        return True



def findline(event,triadsets):
    # lines = open('E:\SIGIR\SVM\特征\反转/' + event + '/' + event + '.txt', 'r').readlines()
    # for line in lines:
    #     line = line.replace('\n','')
    #     isresult = True
    #     for item in triadset:
    #         if item not in line:
    #             isresult = False
    #     if isresult == True:
    #         return line
    resultlines = []
    lines = open('E:\SIGIR\SVM\特征\反转/' + event + '/' + event + '.txt', 'r').readlines()
    for triadset in triadsets:
        for line in lines:
            line = line.replace('\n', '')
            isresult = True
            for item in triadset:
                if item not in line:
                    isresult = False
            if isresult == True and line not in resultlines:
                resultlines.append(line)
                break
    return resultlines

def findreverse(event):
    sentiment_words = []
    sentiment_dict = {}
    not_words = [w.strip() for w in open('D:/success/ciku/notDict.txt').readlines()]
    lines = open('D:\success\情感词典\BosonNLP_sentiment_score\BosonNLP_sentiment_score.txt', 'r',
                 encoding='UTF-8').readlines()
    for line in lines:
        word, score = line.split()
        sentiment_dict[word] = float(score)
        sentiment_words.append(word)
    lines = open('E:\SIGIR\SVM\特征\反转/' + event + '/' + event + '三元组集合.txt', 'r', encoding='utf-8').readlines()
    triadsets = []
    for line in lines:
        triadset = line.split()
        triadsets.append(triadset)
    binlist = []
    for triadset in triadsets:
        binlist.append(triadset[2])
    repeatbin = []
    for i in binlist:
        if binlist.count(i) >= 2 and i not in repeatbin and judgeoneline(event, i, triadsets):
            repeatbin.append(i)
    #print(repeatbin)
    # print(judgeoneline(event,'电话',triadsets))
    triadsets2 = []
    for triadset in triadsets:
        if triadset[2] in repeatbin:
            triadsets2.append(triadset)
    sentiment = []
    #print(triadsets2)
    for triadset in triadsets2:
        if triadset[1] in sentiment_words:
            if sentiment_dict[triadset[1]] > 0:
                sentiment.append('P')
            else:
                sentiment.append('N')
    #print(sentiment)
    notsum = []
    # for item in triadsets2:
    #     lineset.append(findline(event, item))
    lineset = findline(event,triadsets2)
    for i in range(0, len(lineset)):
        sumnot = 0
        for word in not_words:
            if word in lineset[i]:
                sumnot+=1
        notsum.append(sumnot)
    #print(lineset)
    #print(notsum)
    for i in range(0, len(sentiment)):
        if notsum[i] % 2 == 1:
            if sentiment[i] == 'P':
                sentiment[i] = 'N'
            else:
                sentiment[i] = 'P'
    #print(sentiment)
    reverseindex = []
    for i in range(1,len(sentiment)):
        if sentiment[i] != sentiment[0]:
            reverseindex.append(i)
    resultlines = []
    for i in reverseindex:
        resultlines.append(lineset[i])
    reversewords = []
    lines = open('E:\SIGIR\SVM\特征\反转\反转词汇.txt','r').readlines()
    for line in lines:
        line = line.replace('\n','')
        reversewords.append(line)
    #print(reversewords)
    lines = open('E:\SIGIR\SVM\特征\反转/' + event + '/' + event + '.txt', 'r').readlines()
    for line in lines:
        line = line.replace('\n', '')
        isresult = False
        for word in reversewords:
            if word in line:
                isresult = True

        if isresult == True and line not in resultlines:
            resultlines.append(line)
    return resultlines

def Chusquare(eventlist):
    lines = []
    stop_words = [w.strip() for w in open('D:/success/ciku/chinese_stopword.txt').readlines()]
    for event in eventlist:
        lines = lines + open('E:\SIGIR\事件新闻/'+event+'.txt','r').readlines()
    sentimentindex = {}
    sentimentindex['产妇'] = [0,1,4,6,8,15,16]
    sentimentindex['红黄蓝'] = [4,9,13,14]
    sentimentindex['山东'] = [2,6,8]
    sentimentindex['魏则西'] = [2,5,8,9]
    sentimentlist = []
    for event in eventlist:
        lines1 = open('E:\SIGIR\事件新闻/' + event + '.txt', 'r').readlines()
        for i in sentimentindex[event]:
            sentimentlist.append(lines1[i])
    N = len(lines)
    wordset = []
    for line in lines:
        wordlist = jieba.cut(line)
        wordlist = [w for w in wordlist if w.replace('.','').isdigit() is False and len(w)>=2 and w not in stop_words]
        wordset = wordset + wordlist
    wordset2 = []
    for w in wordset:
        if w not in wordset2:
            wordset2.append(w)
    wordset = wordset2
    wordscore = {}
    for word in wordset:
        #print(word)
        a = 0
        b = 0
        c = 0
        d = 0
        for line in lines:
            if word in line and line in sentimentlist:
                a += 1
            if word in line and line not in sentimentlist:
                b += 1
            if word not in line and line in sentimentlist:
                c += 1
            if word not in line and line not in sentimentlist:
                d += 1
        wordscore[word] = (N*pow(a*d-c*b,2))/((a+c)*(b+d)*(a+b)*(c+d))
    linescores = {}
    for line in lines:
        #print(line)
        line = line.replace('\n','')
        sum = 0
        for word in wordset:
            if word in line:
                #print(word+' '+str(wordscore[word]))
                sum = sum + wordscore[word]
        linescores[line] = sum
    return linescores,wordscore

def tf(word,line):
    stop_words = [w.strip() for w in open('D:/success/ciku/chinese_stopword.txt').readlines()]
    wordlist = jieba.cut(line)
    wordlist = [w for w in wordlist if w.replace('.', '').isdigit() is False and len(w) >= 2 and w not in stop_words]
    #print(wordlist)
    s = len(wordlist)
    c = wordlist.count(word)
    return c/s

def idf(word,lines):
    c = 0
    for line in lines:
        if word in line:
            c+=1
    s = len(lines)
    return math.log10(s/c)

def similarword(event):#得到各个事件中和证据相似的词
    similarwords = []
    lines = open('E:\SIGIR\证据类种子词库/'+event+'/'+event+'.txt', 'r').readlines()
    for line in lines:
        line = line.replace('\n', '')
        similarwords.append(str(line))
    return similarwords

def tfidf(event):#TFIDF feature
    lines = open('E:\SIGIR\事件新闻/' + event + '.txt', 'r').readlines()
    words = similarword(event)
    tfidfdir = {}
    for line in lines:
        line = line.replace('\n', '')
        tfidfscore = 0
        for word in words:
            if word in line:
                tfscore = tf(word,line)
                idfscore = idf(word,lines)
                tfidfscore = tfidfscore + tfscore * idfscore
        tfidfdir[line] = tfidfscore
    return tfidfdir



if __name__ == "__main__":
    # print(findmaxsentiment('产妇'))
    # print(findmaxsentiment('红黄蓝'))
    # print(findmaxsentiment('山东'))
    # print(findreverse('产妇'))
    # print(findreverse('红黄蓝'))
    # print(findreverse('魏则西'))
    #linescores,wordscore = Chusquare(['产妇','红黄蓝','山东','魏则西'])
    linescores, wordscore = Chusquare(['魏则西'])
    linescores = sorted(linescores.items(), key=lambda x: x[1], reverse=True)
    # for item in linescores:
    #     print(item[0]+' '+str(item[1]))
    # for item in wordscore.items():
    #     print(item[0]+' '+str(item[1]))

# 特征搞定
# 就算单独的对一个事件进行卡方，效果也是很好的
